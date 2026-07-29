from __future__ import annotations

import csv
import hashlib
import json
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from k12hub.cli import main
from k12hub.generator import (
    ATTENDANCE_STATUSES,
    ERROR_TYPES,
    INSTRUCTIONAL_MINUTES,
    GeneratorOptions,
    generate_data,
)


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as input_file:
        return list(csv.DictReader(input_file))


def _json_lines(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]


def _xlsx_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["assessments"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def test_same_seed_produces_identical_checksums(tmp_path: Path) -> None:
    first = generate_data(
        GeneratorOptions(seed=2026, students=12, output_directory=tmp_path / "first")
    )
    second = generate_data(
        GeneratorOptions(seed=2026, students=12, output_directory=tmp_path / "second")
    )

    assert first.run_id == second.run_id
    assert first.manifest["checksums"] == second.manifest["checksums"]
    with zipfile.ZipFile(first.output_path / "assessments.xlsx") as workbook:
        core_properties = workbook.read("docProps/core.xml")
    assert core_properties.count(b"2000-01-01T00:00:00Z") == 2


def test_different_seed_changes_checksums(tmp_path: Path) -> None:
    first = generate_data(
        GeneratorOptions(seed=2026, students=12, output_directory=tmp_path / "first")
    )
    second = generate_data(
        GeneratorOptions(seed=2027, students=12, output_directory=tmp_path / "second")
    )

    assert first.manifest["checksums"] != second.manifest["checksums"]


def test_clean_dataset_has_consistent_relationships(tmp_path: Path) -> None:
    result = generate_data(GeneratorOptions(seed=2026, students=18, output_directory=tmp_path))
    students = _csv_rows(result.output_path / "students.csv")
    enrollments = _csv_rows(result.output_path / "enrollments.csv")
    attendance = _json_lines(result.output_path / "attendance_events.jsonl")
    assessments = _xlsx_rows(result.output_path / "assessments.xlsx")

    student_ids = {row["student_id"] for row in students}
    schools_by_student = {row["student_id"]: row["school_id"] for row in students}
    enrollment_by_student = {row["student_id"]: row for row in enrollments}
    assert len(student_ids) == len(students)
    assert all(student_ids)
    assert {row["student_id"] for row in enrollments} == student_ids

    for row in attendance:
        student_id = row["student_id"]
        enrollment = enrollment_by_student[student_id]
        exit_date = enrollment["exit_date"] or "9999-12-31"
        assert enrollment["entry_date"] <= row["instructional_date"] <= exit_date
        assert row["school_id"] == schools_by_student[student_id]
        assert row["attendance_status"] in ATTENDANCE_STATUSES
        assert 0 <= row["minutes_attended"] <= INSTRUCTIONAL_MINUTES

    for row in assessments:
        assert row["student_id"] in student_ids
        assert row["school_id"] == schools_by_student[row["student_id"]]


def test_all_enabled_errors_are_present(tmp_path: Path) -> None:
    result = generate_data(
        GeneratorOptions(
            seed=2026,
            students=16,
            error_rate=0.1,
            enabled_error_types=ERROR_TYPES,
            output_directory=tmp_path,
        )
    )
    students = _csv_rows(result.output_path / "students.csv")
    enrollments = _csv_rows(result.output_path / "enrollments.csv")
    attendance = _json_lines(result.output_path / "attendance_events.jsonl")
    assessments = _xlsx_rows(result.output_path / "assessments.xlsx")
    valid_student_ids = {row["student_id"] for row in students if row["student_id"]}

    assert all(count > 0 for count in result.manifest["injected_errors"].values())
    student_id_counts = Counter(row["student_id"] for row in students if row["student_id"])
    assert any(count > 1 for count in student_id_counts.values())
    assert any(not row["student_id"] for row in students)
    assert any(row["grade_level"] == "99" for row in students)
    assert "unexpected_schema_drift" in students[0] or any(
        row.get("unexpected_schema_drift") == "injected" for row in students
    )
    assert any(row["school_id"] == "SYN-UNKNOWN-SCHOOL" for row in enrollments)
    assert len(enrollments) > result.manifest["baseline_record_counts"]["enrollments.csv"]
    assert any(row["student_id"] not in valid_student_ids for row in attendance)
    attendance_keys = Counter(
        (row["student_id"], row["school_id"], row["instructional_date"]) for row in attendance
    )
    assert any(count > 1 for count in attendance_keys.values())
    primary_enrollments = {
        row["student_id"]: row
        for row in enrollments
        if not row["enrollment_id"].endswith("-OVERLAP")
    }
    assert any(
        row["student_id"] in primary_enrollments
        and (
            row["instructional_date"] < primary_enrollments[row["student_id"]]["entry_date"]
            or (
                primary_enrollments[row["student_id"]]["exit_date"]
                and row["instructional_date"] > primary_enrollments[row["student_id"]]["exit_date"]
            )
        )
        for row in attendance
    )
    assert any(row["attendance_status"] == "invalid_status" for row in attendance)
    assert any(row["minutes_attended"] > INSTRUCTIONAL_MINUTES for row in attendance)
    assert any(float(row["scale_score"]) > 100 for row in assessments)


def test_manifest_counts_match_generated_files(tmp_path: Path) -> None:
    result = generate_data(GeneratorOptions(seed=2026, students=10, output_directory=tmp_path))
    counts = result.manifest["record_counts"]

    assert counts["students.csv"] == len(_csv_rows(result.output_path / "students.csv"))
    assert counts["enrollments.csv"] == len(_csv_rows(result.output_path / "enrollments.csv"))
    assert counts["attendance_events.jsonl"] == len(
        _json_lines(result.output_path / "attendance_events.jsonl")
    )
    assert counts["assessments.xlsx"] == len(_xlsx_rows(result.output_path / "assessments.xlsx"))
    assert set(result.manifest["checksums"]) == {
        "assessments.xlsx",
        "attendance_events.jsonl",
        "enrollments.csv",
        "expected_metrics.json",
        "students.csv",
    }
    for filename, expected_checksum in result.manifest["checksums"].items():
        actual_checksum = hashlib.sha256((result.output_path / filename).read_bytes()).hexdigest()
        assert actual_checksum == expected_checksum


def test_expected_metrics_match_generated_clean_records(tmp_path: Path) -> None:
    result = generate_data(GeneratorOptions(seed=2026, students=14, output_directory=tmp_path))
    metrics = result.expected_metrics["metrics"]
    attendance = _json_lines(result.output_path / "attendance_events.jsonl")
    assessments = _xlsx_rows(result.output_path / "assessments.xlsx")

    expected_attendance_rate = round(
        sum(row["minutes_attended"] for row in attendance)
        / (len(attendance) * INSTRUCTIONAL_MINUTES),
        6,
    )
    attendance_by_student: dict[str, list[dict[str, Any]]] = {}
    for row in attendance:
        attendance_by_student.setdefault(row["student_id"], []).append(row)
    chronic_count = sum(
        (
            sum(INSTRUCTIONAL_MINUTES - row["minutes_attended"] for row in student_rows)
            / (len(student_rows) * INSTRUCTIONAL_MINUTES)
        )
        >= 0.10
        for student_rows in attendance_by_student.values()
    )
    proficient = sum(
        row["performance_level"] in {"meets_standard", "exceeds_standard"} for row in assessments
    )
    assert metrics["student_count"] == 14
    assert metrics["enrollment_count"] == 14
    assert metrics["assessment_event_count"] == len(assessments)
    assert metrics["attendance_rate"] == expected_attendance_rate
    assert metrics["chronic_absence_rate"] == round(
        chronic_count / len(attendance_by_student),
        6,
    )
    assert metrics["assessment_proficiency_rate"] == round(proficient / len(assessments), 6)


def test_generate_data_cli_writes_small_dataset(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    exit_code = main(
        [
            "generate-data",
            "--seed",
            "2026",
            "--students",
            "5",
            "--school-year",
            "2025-2026",
            "--output-directory",
            str(tmp_path),
        ]
    )

    assert exit_code == 0
    assert "Synthetic dataset generated:" in capsys.readouterr().out
    assert list(tmp_path.glob("run-*/generation_manifest.json"))
