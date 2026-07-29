"""Contract-driven parsing and transactional loading of raw MinIO objects."""

from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from fnmatch import fnmatch
from typing import Any, Protocol
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import Connection, Engine, text

from k12hub.config import MinioSettings
from k12hub.contracts import (
    ConfigurationBundle,
    ContractField,
    DataType,
    FileFormat,
    SourceContract,
)
from k12hub.database import transaction
from k12hub.object_store import ObjectStorageClient

BATCH_SIZE = 500
PARSING_RULE_CODE = "source_contract_parse_error"
PARSING_RULE_ID = UUID("00000000-0000-0000-0000-000000000700")
NORMALIZED_COLUMN_PATTERN = re.compile(r"[^a-z0-9]+")

# Contract configuration selects only from this allowlist. Identifiers never come from file data.
STAGING_BUSINESS_COLUMNS: dict[str, tuple[str, ...]] = {
    "staging.sis_student": (
        "student_id",
        "local_student_number",
        "first_name",
        "last_name",
        "birth_date",
        "gender",
        "grade_level",
        "district_id",
        "school_id",
        "active",
    ),
    "staging.sis_enrollment": (
        "enrollment_id",
        "student_id",
        "district_id",
        "school_id",
        "academic_year",
        "grade_level",
        "entry_date",
        "exit_date",
        "enrollment_status",
    ),
    "staging.attendance_event": (
        "student_id",
        "district_id",
        "school_id",
        "instructional_date",
        "attendance_status",
        "minutes_attended",
        "reason_code",
        "recorded_at",
    ),
    "staging.assessment_event": (
        "assessment_event_id",
        "student_id",
        "district_id",
        "school_id",
        "academic_year",
        "assessment_name",
        "subject",
        "assessment_date",
        "scale_score",
        "performance_level",
    ),
}


class StagingError(RuntimeError):
    """Base error for staging operations."""


class FileStructureError(StagingError):
    """Raised when an object does not have the contract's required structure."""


class StagingLoadError(StagingError):
    """Raised when a staging pipeline run cannot be loaded."""


@dataclass(frozen=True)
class SourceFileReference:
    """Audit metadata needed to retrieve and stage one loaded raw file."""

    pipeline_run_id: UUID
    source_file_id: UUID
    source_system: str
    original_filename: str
    object_path: str


@dataclass(frozen=True)
class ParsedRow:
    """One valid source row with raw and typed representations."""

    source_row_number: int
    raw_payload: dict[str, Any]
    business_values: dict[str, Any]


@dataclass(frozen=True)
class RejectedRow:
    """One source row that could not be parsed against its contract."""

    source_row_number: int
    raw_payload: dict[str, Any]
    error_message: str


@dataclass(frozen=True)
class ParsedFile:
    """All row-level parse outcomes for one immutable object."""

    discovered: int
    parsed_rows: tuple[ParsedRow, ...]
    rejected_rows: tuple[RejectedRow, ...]

    @property
    def parsed(self) -> int:
        return len(self.parsed_rows)

    @property
    def rejected(self) -> int:
        return len(self.rejected_rows)


@dataclass(frozen=True)
class RowCountReconciliation:
    """Reconciled counts for one staged source file."""

    source_file_id: UUID
    discovered: int
    parsed: int
    loaded: int
    rejected: int
    status: str


@dataclass(frozen=True)
class StagingResult:
    """Aggregate counts returned by a staging load command."""

    pipeline_run_id: UUID
    files: int
    discovered: int
    parsed: int
    loaded: int
    rejected: int


class StagingMetadataStore(Protocol):
    """Database behavior used by the staging service."""

    def list_loaded_source_files(self, pipeline_run_id: UUID) -> list[SourceFileReference]:
        """Return raw objects belonging to one ingestion pipeline run."""

    def load_file(
        self,
        source_file: SourceFileReference,
        contract: SourceContract,
        parsed_file: ParsedFile,
    ) -> RowCountReconciliation:
        """Atomically batch-load valid and rejected rows and reconcile counts."""


def normalize_column_name(value: object) -> str:
    """Normalize a source column name to stable lower snake case."""

    normalized = NORMALIZED_COLUMN_PATTERN.sub("_", str(value).strip().lower()).strip("_")
    if not normalized:
        raise FileStructureError(f"Column name {value!r} normalizes to an empty value")
    return normalized


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _json_payload(values: Mapping[str, Any]) -> dict[str, Any]:
    return {str(key): _json_value(value) for key, value in values.items()}


def _validate_headers(headers: Sequence[object], contract: SourceContract) -> list[str]:
    normalized = [normalize_column_name(header) for header in headers]
    if len(normalized) != len(set(normalized)):
        raise FileStructureError("Source columns collide after normalization")
    missing = sorted(set(contract.required_fields) - set(normalized))
    if missing:
        raise FileStructureError(f"Missing required columns: {', '.join(missing)}")
    return normalized


def _parse_boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "1"}:
        return True
    if normalized in {"false", "0"}:
        return False
    raise ValueError(f"expected boolean, received {value!r}")


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))


def _parse_field(value: Any, field: ContractField) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if field.data_type is DataType.STRING:
        parsed: Any = str(value).strip()
    elif field.data_type is DataType.INTEGER:
        if isinstance(value, bool):
            raise ValueError(f"expected integer, received {value!r}")
        parsed = int(str(value).strip())
    elif field.data_type is DataType.DECIMAL:
        try:
            parsed = Decimal(str(value).strip())
        except InvalidOperation as error:
            raise ValueError(f"expected decimal, received {value!r}") from error
    elif field.data_type is DataType.BOOLEAN:
        parsed = _parse_boolean(value)
    elif field.data_type is DataType.DATE:
        parsed = _parse_date(value)
    elif field.data_type is DataType.DATETIME:
        parsed = _parse_datetime(value)
    else:  # pragma: no cover - the validated enum makes this unreachable
        raise ValueError(f"unsupported data type {field.data_type}")

    if field.accepted_values is not None and str(parsed) not in field.accepted_values:
        raise ValueError(f"{field.name} value {parsed!r} is not accepted")
    return parsed


def _parse_record(
    normalized_record: Mapping[str, Any],
    raw_payload: dict[str, Any],
    row_number: int,
    contract: SourceContract,
) -> ParsedRow | RejectedRow:
    try:
        missing_values = [
            field_name
            for field_name in contract.required_fields
            if normalized_record.get(field_name) is None
            or (
                isinstance(normalized_record.get(field_name), str)
                and not str(normalized_record[field_name]).strip()
            )
        ]
        if missing_values:
            raise ValueError(f"required values missing: {', '.join(missing_values)}")

        business_values = {
            field.name: _parse_field(normalized_record.get(field.name), field)
            for field in contract.expected_fields
        }
        return ParsedRow(
            source_row_number=row_number,
            raw_payload=raw_payload,
            business_values=business_values,
        )
    except (TypeError, ValueError) as error:
        return RejectedRow(
            source_row_number=row_number,
            raw_payload=raw_payload,
            error_message=str(error),
        )


def _parse_tabular_rows(
    headers: Sequence[object],
    rows: Iterable[Sequence[Any]],
    contract: SourceContract,
) -> ParsedFile:
    normalized_headers = _validate_headers(headers, contract)
    original_headers = [str(header) for header in headers]
    parsed_rows: list[ParsedRow] = []
    rejected_rows: list[RejectedRow] = []
    discovered = 0
    for row_number, values in enumerate(rows, start=2):
        discovered += 1
        if len(values) != len(normalized_headers):
            outcome: ParsedRow | RejectedRow = RejectedRow(
                source_row_number=row_number,
                raw_payload={
                    "_headers": original_headers,
                    "_values": [_json_value(v) for v in values],
                },
                error_message=(
                    f"expected {len(normalized_headers)} columns, received {len(values)}"
                ),
            )
        else:
            raw = _json_payload(dict(zip(original_headers, values)))
            normalized_record = dict(zip(normalized_headers, values))
            outcome = _parse_record(normalized_record, raw, row_number, contract)
        if isinstance(outcome, ParsedRow):
            parsed_rows.append(outcome)
        else:
            rejected_rows.append(outcome)
    return ParsedFile(discovered, tuple(parsed_rows), tuple(rejected_rows))


def _parse_csv(content: bytes, contract: SourceContract) -> ParsedFile:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise FileStructureError("CSV object is not valid UTF-8") from error
    reader = csv.reader(io.StringIO(decoded, newline=""), strict=True)
    try:
        headers = next(reader)
        return _parse_tabular_rows(headers, reader, contract)
    except StopIteration as error:
        raise FileStructureError("CSV object is empty") from error
    except csv.Error as error:
        raise FileStructureError(f"CSV structure is invalid: {error}") from error


def _parse_json_lines(content: bytes, contract: SourceContract) -> ParsedFile:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise FileStructureError("JSON Lines object is not valid UTF-8") from error

    parsed_rows: list[ParsedRow] = []
    rejected_rows: list[RejectedRow] = []
    discovered = 0
    required_columns_validated = False
    for row_number, line in enumerate(decoded.splitlines(), start=1):
        discovered += 1
        try:
            raw_value = json.loads(line)
            if not isinstance(raw_value, dict):
                raise ValueError("JSON Lines row must be an object")
            original_headers = list(raw_value)
            normalized_headers = _validate_headers(original_headers, contract)
            required_columns_validated = True
            normalized_record = dict(zip(normalized_headers, raw_value.values()))
            outcome = _parse_record(
                normalized_record,
                _json_payload(raw_value),
                row_number,
                contract,
            )
        except (json.JSONDecodeError, ValueError, FileStructureError) as error:
            outcome = RejectedRow(
                source_row_number=row_number,
                raw_payload={"_raw_line": line},
                error_message=str(error),
            )
        if isinstance(outcome, ParsedRow):
            parsed_rows.append(outcome)
        else:
            rejected_rows.append(outcome)

    if discovered == 0:
        raise FileStructureError("JSON Lines object is empty")
    if not required_columns_validated:
        raise FileStructureError("No JSON Lines row contains the required columns")
    return ParsedFile(discovered, tuple(parsed_rows), tuple(rejected_rows))


def _parse_xlsx(content: bytes, contract: SourceContract) -> ParsedFile:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise FileStructureError("XLSX object could not be opened") from error
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise FileStructureError("XLSX object does not contain an active worksheet")
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as error:
            raise FileStructureError("XLSX object is empty") from error
        return _parse_tabular_rows(headers, rows, contract)
    finally:
        workbook.close()


def parse_object(content: bytes, contract: SourceContract) -> ParsedFile:
    """Parse one MinIO object according to its validated source contract."""

    if contract.file_format is FileFormat.CSV:
        return _parse_csv(content, contract)
    if contract.file_format is FileFormat.JSON_LINES:
        return _parse_json_lines(content, contract)
    if contract.file_format is FileFormat.XLSX:
        return _parse_xlsx(content, contract)
    raise FileStructureError(f"Unsupported file format: {contract.file_format}")


def _contract_for_file(
    source_file: SourceFileReference,
    configuration: ConfigurationBundle,
) -> SourceContract:
    matches = [
        contract
        for contract in configuration.contracts.values()
        if contract.source_system == source_file.source_system
        and fnmatch(source_file.original_filename, contract.file_pattern)
    ]
    if len(matches) != 1:
        raise StagingLoadError(
            f"{source_file.original_filename} matched {len(matches)} contracts; "
            "expected exactly one"
        )
    contract = matches[0]
    if contract.destination_staging_table not in STAGING_BUSINESS_COLUMNS:
        raise StagingLoadError(
            f"Contract destination is not allowlisted: {contract.destination_staging_table}"
        )
    return contract


class PostgresStagingMetadataStore:
    """PostgreSQL staging loader with one atomic transaction per source file."""

    def __init__(self, engine: Engine, *, batch_size: int = BATCH_SIZE) -> None:
        if batch_size < 1:
            raise ValueError("batch_size must be positive")
        self._engine = engine
        self._batch_size = batch_size

    def list_loaded_source_files(self, pipeline_run_id: UUID) -> list[SourceFileReference]:
        with transaction(engine=self._engine) as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT pipeline_run_id, source_file_id, source_system,
                           original_filename, object_path
                    FROM audit.source_file
                    WHERE pipeline_run_id = :pipeline_run_id
                      AND status = 'loaded'
                    ORDER BY source_file_id
                    """
                ),
                {"pipeline_run_id": pipeline_run_id},
            ).mappings()
            return [
                SourceFileReference(
                    pipeline_run_id=UUID(str(row["pipeline_run_id"])),
                    source_file_id=UUID(str(row["source_file_id"])),
                    source_system=str(row["source_system"]),
                    original_filename=str(row["original_filename"]),
                    object_path=str(row["object_path"]),
                )
                for row in rows
            ]

    def _register_parsing_rule(self, connection: Connection) -> UUID:
        rule_id = connection.execute(
            text(
                """
                INSERT INTO metadata.data_quality_rule (
                    rule_id, rule_code, name, description, rule_type,
                    severity, configuration
                )
                VALUES (
                    :rule_id, :rule_code, :name, :description, :rule_type,
                    :severity, CAST(:configuration AS JSONB)
                )
                ON CONFLICT (rule_code) DO UPDATE
                SET updated_at = CURRENT_TIMESTAMP
                RETURNING rule_id
                """
            ),
            {
                "rule_id": PARSING_RULE_ID,
                "rule_code": PARSING_RULE_CODE,
                "name": "Source contract parse error",
                "description": "A raw source row could not be parsed against its contract.",
                "rule_type": "required",
                "severity": "error",
                "configuration": "{}",
            },
        ).scalar_one()
        return UUID(str(rule_id))

    def _insert_batches(
        self,
        connection: Connection,
        statement: str,
        rows: Sequence[dict[str, Any]],
    ) -> None:
        for offset in range(0, len(rows), self._batch_size):
            connection.execute(text(statement), rows[offset : offset + self._batch_size])

    def _before_reconciliation(
        self,
        connection: Connection,
        source_file: SourceFileReference,
    ) -> None:
        """Test seam executed inside the file transaction before reconciliation."""

    def load_file(
        self,
        source_file: SourceFileReference,
        contract: SourceContract,
        parsed_file: ParsedFile,
    ) -> RowCountReconciliation:
        destination = contract.destination_staging_table
        business_columns = STAGING_BUSINESS_COLUMNS[destination]
        table_name = destination.removeprefix("staging.")
        shared_columns = (
            "pipeline_run_id",
            "source_file_id",
            "source_row_number",
            "source_schema_version",
            "raw_payload",
        )
        insert_columns = (*shared_columns, *business_columns)
        placeholders = [
            "CAST(:raw_payload AS JSONB)" if column == "raw_payload" else f":{column}"
            for column in insert_columns
        ]
        staging_statement = (
            f"INSERT INTO staging.{table_name} ({', '.join(insert_columns)}) "
            f"VALUES ({', '.join(placeholders)}) "
            "ON CONFLICT (source_file_id, source_row_number) DO NOTHING"
        )
        staging_rows = [
            {
                "pipeline_run_id": source_file.pipeline_run_id,
                "source_file_id": source_file.source_file_id,
                "source_row_number": row.source_row_number,
                "source_schema_version": contract.schema_version,
                "raw_payload": json.dumps(row.raw_payload, sort_keys=True),
                **row.business_values,
            }
            for row in parsed_file.parsed_rows
        ]

        with transaction(engine=self._engine) as connection:
            locked = connection.execute(
                text(
                    """
                    SELECT 1
                    FROM audit.source_file
                    WHERE source_file_id = :source_file_id
                      AND pipeline_run_id = :pipeline_run_id
                      AND status = 'loaded'
                    FOR UPDATE
                    """
                ),
                {
                    "source_file_id": source_file.source_file_id,
                    "pipeline_run_id": source_file.pipeline_run_id,
                },
            ).first()
            if locked is None:
                raise StagingLoadError(f"Source file is unavailable: {source_file.source_file_id}")

            parse_rule_id = self._register_parsing_rule(connection)
            self._insert_batches(connection, staging_statement, staging_rows)
            rejected_rows = [
                {
                    "pipeline_run_id": source_file.pipeline_run_id,
                    "source_file_id": source_file.source_file_id,
                    "source_system": source_file.source_system,
                    "source_row_number": row.source_row_number,
                    "rule_id": parse_rule_id,
                    "error_message": row.error_message,
                    "raw_payload": json.dumps(row.raw_payload, sort_keys=True),
                }
                for row in parsed_file.rejected_rows
            ]
            self._insert_batches(
                connection,
                """
                INSERT INTO quarantine.rejected_record (
                    pipeline_run_id, source_file_id, source_system, source_row_number,
                    rule_id, severity, error_message, raw_payload
                )
                VALUES (
                    :pipeline_run_id, :source_file_id, :source_system, :source_row_number,
                    :rule_id, 'error', :error_message, CAST(:raw_payload AS JSONB)
                )
                ON CONFLICT (source_file_id, source_row_number, rule_id) DO NOTHING
                """,
                rejected_rows,
            )

            loaded = int(
                connection.execute(
                    text(f"SELECT count(*) FROM staging.{table_name} WHERE source_file_id = :id"),
                    {"id": source_file.source_file_id},
                ).scalar_one()
            )
            rejected = int(
                connection.execute(
                    text(
                        """
                        SELECT count(*)
                        FROM quarantine.rejected_record
                        WHERE source_file_id = :source_file_id
                          AND rule_id = :rule_id
                        """
                    ),
                    {
                        "source_file_id": source_file.source_file_id,
                        "rule_id": parse_rule_id,
                    },
                ).scalar_one()
            )
            self._before_reconciliation(connection, source_file)
            balanced = (
                parsed_file.discovered == parsed_file.parsed + rejected
                and loaded == parsed_file.parsed
            )
            if not balanced:
                raise StagingLoadError(
                    f"Row counts did not reconcile for source file {source_file.source_file_id}"
                )
            status = "matched"
            actual = loaded + rejected
            connection.execute(
                text(
                    """
                    INSERT INTO audit.row_count_reconciliation (
                        pipeline_run_id, source_file_id, source_system, stage_name,
                        expected_row_count, actual_row_count, difference, status,
                        discovered_row_count, parsed_row_count, loaded_row_count,
                        rejected_row_count
                    )
                    VALUES (
                        :pipeline_run_id, :source_file_id, :source_system, :stage_name,
                        :expected, :actual, :difference, :status,
                        :discovered, :parsed, :loaded, :rejected
                    )
                    ON CONFLICT (source_file_id, stage_name) DO UPDATE
                    SET expected_row_count = EXCLUDED.expected_row_count,
                        actual_row_count = EXCLUDED.actual_row_count,
                        difference = EXCLUDED.difference,
                        status = EXCLUDED.status,
                        discovered_row_count = EXCLUDED.discovered_row_count,
                        parsed_row_count = EXCLUDED.parsed_row_count,
                        loaded_row_count = EXCLUDED.loaded_row_count,
                        rejected_row_count = EXCLUDED.rejected_row_count,
                        reconciled_at = CURRENT_TIMESTAMP,
                        updated_at = CURRENT_TIMESTAMP
                    """
                ),
                {
                    "pipeline_run_id": source_file.pipeline_run_id,
                    "source_file_id": source_file.source_file_id,
                    "source_system": source_file.source_system,
                    "stage_name": table_name,
                    "expected": parsed_file.discovered,
                    "actual": actual,
                    "difference": parsed_file.discovered - actual,
                    "status": status,
                    "discovered": parsed_file.discovered,
                    "parsed": parsed_file.parsed,
                    "loaded": loaded,
                    "rejected": rejected,
                },
            )
            connection.execute(
                text(
                    """
                    UPDATE audit.source_file
                    SET row_count = :row_count, updated_at = CURRENT_TIMESTAMP
                    WHERE source_file_id = :source_file_id
                    """
                ),
                {
                    "row_count": parsed_file.discovered,
                    "source_file_id": source_file.source_file_id,
                },
            )

        return RowCountReconciliation(
            source_file_id=source_file.source_file_id,
            discovered=parsed_file.discovered,
            parsed=parsed_file.parsed,
            loaded=loaded,
            rejected=rejected,
            status=status,
        )


class StagingLoaderService:
    """Retrieve loaded raw objects from MinIO, parse them, and stage their rows."""

    def __init__(
        self,
        metadata_store: StagingMetadataStore,
        object_store: ObjectStorageClient,
        minio_settings: MinioSettings,
    ) -> None:
        self._metadata_store = metadata_store
        self._object_store = object_store
        self._raw_bucket = minio_settings.raw_bucket

    def load(
        self,
        pipeline_run_id: UUID,
        configuration: ConfigurationBundle,
    ) -> StagingResult:
        """Load all successfully ingested source files for one pipeline run."""

        source_files = self._metadata_store.list_loaded_source_files(pipeline_run_id)
        if not source_files:
            raise StagingLoadError(
                f"No loaded source files found for pipeline run {pipeline_run_id}"
            )

        reconciliations: list[RowCountReconciliation] = []
        try:
            for source_file in source_files:
                contract = _contract_for_file(source_file, configuration)
                content = self._object_store.read_object(
                    self._raw_bucket,
                    source_file.object_path,
                )
                parsed_file = parse_object(content, contract)
                reconciliations.append(
                    self._metadata_store.load_file(source_file, contract, parsed_file)
                )
        except StagingError:
            raise
        except Exception as error:
            raise StagingLoadError(
                f"Staging load failed for pipeline run {pipeline_run_id}: {error}"
            ) from error

        return StagingResult(
            pipeline_run_id=pipeline_run_id,
            files=len(reconciliations),
            discovered=sum(item.discovered for item in reconciliations),
            parsed=sum(item.parsed for item in reconciliations),
            loaded=sum(item.loaded for item in reconciliations),
            rejected=sum(item.rejected for item in reconciliations),
        )
